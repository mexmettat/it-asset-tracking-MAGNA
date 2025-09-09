from datetime import datetime
import os
import io
import pandas as pd
from flask import Flask, render_template, request, redirect, url_for, flash
from flask_sqlalchemy import SQLAlchemy
from collections import defaultdict
from sqlalchemy import func
from flask import send_file
from datetime import datetime
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# --- PDF İÇİN FONT ---
BASE_DIR  = os.path.dirname(__file__)
FONTS_DIR = os.path.join(BASE_DIR, "static", "fonts")

MONTS_REG  = os.path.join(FONTS_DIR, "Montserrat-Regular.ttf")
MONTS_BOLD = os.path.join(FONTS_DIR, "Montserrat-BoldItalic.ttf") 

# --- App & DB ---
app = Flask(__name__)
app.config["SECRET_KEY"] = os.getenv("SECRET_KEY", "dev-secret-change-me")
app.config["SQLALCHEMY_DATABASE_URI"] = os.getenv("SQLALCHEMY_DATABASE_URI", "sqlite:///app.db")
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db = SQLAlchemy(app)

# --- Models ---
class Device(db.Model):
    __tablename__ = "devices"
    id = db.Column(db.Integer, primary_key=True)
    brand = db.Column(db.String(80), nullable=False)
    type = db.Column(db.String(40), nullable=False)   # Monitor, Laptop PC vb.
    serial = db.Column(db.String(120), unique=True, nullable=False)
    home_code = db.Column(db.String(80))
    os = db.Column(db.String(80))
    quantity = db.Column(db.Integer, default=1, nullable=False)
    is_faulty = db.Column(db.Boolean, default=False, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)

     # İlişkiler: hepsi back_populates ile
    service_records = db.relationship(
        "ServiceRecord", back_populates="device", cascade="all, delete-orphan"
    )
    faults = db.relationship(
        "Fault", back_populates="device", cascade="all, delete-orphan"
    )
    usages = db.relationship(
        "Usage", back_populates="device", cascade="all, delete-orphan"
    )

class ServiceRecord(db.Model):
    __tablename__ = "service_records"
    id = db.Column(db.Integer, primary_key=True)
    device_id = db.Column(db.Integer, db.ForeignKey("devices.id"), nullable=False)
    service_date = db.Column(db.Date, nullable=False)
    status = db.Column(db.String(30), default="Beklemede")  # Beklemede / Gönderildi / Geldi / Tamir Edildi / Teslim Edildi
    description = db.Column(db.Text)
    cost = db.Column(db.Float, default=0.0)
    return_date = db.Column(db.Date)  # ⇠ servisten dönüş tarihi
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    device = db.relationship("Device", back_populates="service_records")

class Parameter(db.Model):
    __tablename__ = "parameters"
    id = db.Column(db.Integer, primary_key=True)
    key = db.Column(db.String(120), unique=True, nullable=False)
    value = db.Column(db.Text, default="")

class Fault(db.Model):
    __tablename__ = "faults"
    id = db.Column(db.Integer, primary_key=True)
    device_id = db.Column(db.Integer, db.ForeignKey("devices.id"), nullable=False)
    title = db.Column(db.String(255), nullable=False)
    is_open = db.Column(db.Boolean, default=True, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    resolved_at = db.Column(db.DateTime)
    device = db.relationship("Device", back_populates="faults")

class Usage(db.Model):
    __tablename__ = "usages"
    id = db.Column(db.Integer, primary_key=True)
    device_id = db.Column(db.Integer, db.ForeignKey("devices.id"), nullable=False)
    user_name = db.Column(db.String(120), nullable=False)
    assigned_date = db.Column(db.Date, default=datetime.utcnow)
    return_date = db.Column(db.Date)
    is_active = db.Column(db.Boolean, default=True, nullable=False)
    device = db.relationship("Device", back_populates="usages")

# --- One-time init ---
with app.app_context():
    db.create_all()

# --- Helpers ---
def parse_date(date_str):
    try:
        return datetime.strptime(date_str, "%Y-%m-%d").date()
    except Exception:
        return datetime.utcnow().date()

def parse_date_optional(date_str):
    s = (date_str or "").strip()
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except Exception:
        return None

def _split_list(text: str):
    if not text:
        return []
    raw = text.replace("\r", "\n").replace(",", "\n").split("\n")
    return [x.strip() for x in raw if x.strip()]

def get_param_list(key: str):
    p = Parameter.query.filter_by(key=key).first()
    return _split_list(p.value if p else "")

def _open_fault_if_missing(device_id: int, title: str):
    open_fault = Fault.query.filter_by(device_id=device_id, is_open=True).first()
    if not open_fault:
        db.session.add(Fault(device_id=device_id, title=title, is_open=True))

def device_is_locked(device_id: int) -> bool:
    """Cihaz aktif kullanımda mı veya açık serviste mi?"""
    in_use = (
        db.session.query(Usage.id)
        .filter(Usage.device_id == device_id, Usage.is_active.is_(True))
        .first()
        is not None
    )
    in_service = (
        db.session.query(ServiceRecord.id)
        .filter(ServiceRecord.device_id == device_id, ServiceRecord.status.in_(list(OPEN_STATUSES)))
        .first()
        is not None
    )
    return in_use or in_service

def set_param_list(key: str, values: list[str]):
    """Listeyi benzersiz & sıralı tutarak kaydeder (her satıra bir öğe)."""
    clean = sorted(list({v.strip() for v in values if v and v.strip()}), key=str.casefold)
    rec = Parameter.query.filter_by(key=key).first()
    text = "\n".join(clean)
    if rec:
        rec.value = text
    else:
        db.session.add(Parameter(key=key, value=text))
    db.session.commit()

# --- Service status helpers ---
OPEN_STATUSES = {"Beklemede", "Gönderildi", "Geldi", "Tamirde"}
RESOLVED_STATUSES = {"Tamir Edildi", "Teslim Edildi"}

def is_resolved_status(status: str) -> bool:
    """Servis durumu arızanın çözüldüğünü gösteriyorsa True döner."""
    if not status:
        return False
    if status in RESOLVED_STATUSES:
        return True
    s = status.lower()
    return any(k in s for k in ["tamir", "çöz", "onar", "teslim"])

# =============================
#           ROUTES
# =============================

@app.route("/")
def index():
    return redirect(url_for("inventory"))

@app.route("/kullanim")
def kullanim():
    return render_template("kullanim.html")

# -------- Inventory (Depo) --------
@app.route("/inventory")
def inventory():
    q = request.args.get("q", "").strip()
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 10, type=int)

    query = Device.query
    if q:
        like = f"%{q}%"
        query = query.filter(
            db.or_(
                Device.brand.ilike(like),
                Device.type.ilike(like),
                Device.serial.ilike(like),
                Device.home_code.ilike(like),
                Device.os.ilike(like),
            )
        )

    pagination = query.order_by(Device.created_at.desc()).paginate(page=page, per_page=per_page, error_out=False)

    # Parametre listeleri (modal ve edit modalları için)
    brands = get_param_list("brands")        # örn: Dell, HP, Lenovo
    types  = get_param_list("types")         # örn: Laptop PC, Monitor, ...
    oslist = get_param_list("os_list")       # örn: WIN11 23H2, WIN10 21H2 ...

    # aktif kullanımda olan cihazlar
    active_usage_ids = {
        row[0]
        for row in (
            db.session.query(Usage.device_id)
            .filter(Usage.is_active.is_(True))
            .distinct()
            .all()
        )
    }
    open_service_ids = {
        row[0]
        for row in (
            db.session.query(ServiceRecord.device_id)
            .filter(ServiceRecord.status.in_(list(OPEN_STATUSES)))
            .distinct()
            .all()
        )
    }

    return render_template(
        "inventory.html",
        pagination=pagination,
        q=q,
        brands=brands,
        types=types,
        oslist=oslist,
        per_page=per_page,
        active_usage_ids=active_usage_ids,
        open_service_ids=open_service_ids,
    )

@app.route("/inventory/add", methods=["POST"])
def inventory_add():
    brand = request.form.get("brand", "").strip()
    dtype = request.form.get("type", "").strip()
    serial = request.form.get("serial", "").strip()
    home_code = request.form.get("home_code", "").strip()
    os_name = request.form.get("os", "").strip()
    quantity = request.form.get("quantity", type=int, default=1)
    
        # Seri no boşsa "NONE" ver
    if not serial:
        base = "NONE"
        serial = base
        i = 1
        # Çakışma varsa NONE-1, NONE-2, ...
        while Device.query.filter_by(serial=serial).first():
            serial = f"{base}-{i}"
            i += 1

    elif Device.query.filter_by(serial=serial).first():
        flash("Bu seri numarası zaten kayıtlı.", "error")
        return redirect(url_for("inventory"))

    dev = Device(brand=brand, type=dtype, serial=serial, home_code=home_code, os=os_name, quantity=quantity)
    db.session.add(dev)
    db.session.commit()
    flash("Cihaz depoya eklendi.", "ok")
    return redirect(url_for("inventory"))

@app.route("/inventory/update/<int:id>", methods=["POST"])
def inventory_update(id):
    if device_is_locked(id):
        flash("Bu cihaz şu anda kullanımda veya serviste. Düzenleme yapılamaz.", "error")
        return redirect(url_for("inventory"))

    dev = Device.query.get_or_404(id)
    dev.brand = request.form.get("brand", dev.brand).strip()
    dev.type = request.form.get("type", dev.type).strip()
    new_serial = request.form.get("serial", dev.serial).strip()

    if not new_serial:
        # Boş bırakıldıysa -> NONE serisi üret
        base = "NONE"
        candidate = base
        i = 1
        # Kendi dışında çakışma olmasın
        while Device.query.filter(Device.serial == candidate, Device.id != dev.id).first():
            candidate = f"{base}-{i}"
            i += 1
        dev.serial = candidate
    else:
        # Doluysa çakışma kontrolü (kendi dışında)
        exists = Device.query.filter(Device.serial == new_serial, Device.id != dev.id).first()
        if exists:
            flash("Bu seri numarası başka bir cihaza ait.", "error")
            return redirect(url_for("inventory"))
    dev.serial = new_serial
    dev.home_code = request.form.get("home_code", dev.home_code).strip()
    dev.os = request.form.get("os", dev.os).strip()
    dev.quantity = request.form.get("quantity", type=int, default=dev.quantity)
    db.session.commit()
    flash("Cihaz güncellendi.", "ok")
    return redirect(url_for("inventory"))

@app.route("/inventory/delete/<int:id>", methods=["POST"])
def inventory_delete(id):
    if device_is_locked(id):
        flash("Bu cihaz kullanımda veya servisteyken silinemez.", "error")
        return redirect(url_for("inventory"))

    # Fault kayıtlarını sil (ilişkide cascade yoksa FK hatası olmasın)
    Fault.query.filter_by(device_id=id).delete(synchronize_session=False)
    # ServiceRecord'lar cascade ile silinir (modelde tanımlı)
    dev = Device.query.get_or_404(id)
    db.session.delete(dev)
    db.session.commit()
    flash("Cihaz silindi.", "ok")
    return redirect(url_for("inventory"))

@app.route("/inventory/mark_faulty/<int:id>", methods=["POST"])
def inventory_mark_faulty(id):
    if device_is_locked(id):
        flash("Cihaz kullanımda veya servisteyken arızalı işaretlenemez.", "error")
        return redirect(url_for("inventory"))

    dev = Device.query.get_or_404(id)
    dev.is_faulty = True
    # Eğer açık fault yoksa, varsayılan başlıkla bir fault aç
    open_fault = Fault.query.filter_by(device_id=id, is_open=True).first()
    if not open_fault:
        db.session.add(Fault(device_id=id, title=f"Depodan arızalı işaretlendi ({dev.serial})", is_open=True))
    db.session.commit()
    flash("Cihaz arızalı olarak işaretlendi.", "ok")
    return redirect(url_for("inventory"))

@app.route("/inventory/mark_ok/<int:id>", methods=["POST"])
def inventory_mark_ok(id):
    dev = Device.query.get_or_404(id)
    dev.is_faulty = False
    # Açık fault'ları kapat
    for f in Fault.query.filter_by(device_id=id, is_open=True).all():
        f.is_open = False
        f.resolved_at = datetime.utcnow()
    db.session.commit()
    flash("Cihaz normal olarak işaretlendi.", "ok")
    return redirect(url_for("inventory"))

# Adet ±
@app.route("/inventory/qty/<int:id>/<op>", methods=["POST"])
def inventory_qty(id, op):
    dev = Device.query.get_or_404(id)
    if op == "inc":
        dev.quantity = (dev.quantity or 1) + 1
    elif op == "dec":
        dev.quantity = max(1, (dev.quantity or 1) - 1)
    db.session.commit()
    return redirect(request.referrer or url_for("inventory"))

# -------- Faults (Arızalı) --------
@app.route("/faults")
def faults():
    # --- filtreler ---
    q = (request.args.get("q") or "").strip()              # arama (seri/marka/tip/home_code/os)
    page = request.args.get("page", 1, type=int)           # sayfa
    per_page = request.args.get("per_page", 10, type=int)  # sayfa başına kayıt
    brand_sel = (request.args.get("brand") or "").strip()          # YENİ: marka filtresi
    type_sel  = (request.args.get("type") or "").strip()           # YENİ: tip filtresi

    # Sadece arızalı cihazlar
    base = Device.query.filter(Device.is_faulty.is_(True))

    #Marka/Tip filtreleri
    if brand_sel:
        base = base.filter(Device.brand == brand_sel)
    if type_sel:
        base = base.filter(Device.type == type_sel)

    # Arama
    if q:
        like = f"%{q}%"
        base = base.filter(db.or_(
            Device.serial.ilike(like),
            Device.brand.ilike(like),
            Device.type.ilike(like),
            Device.home_code.ilike(like),
            Device.os.ilike(like),
        ))
        # arama yapılınca 1. sayfadan başla
        page = 1

    # Filtre (brand/type) değiştiğinde de 1. sayfadan başla (opsiyonel ama faydalı)
    if request.args.get("brand") is not None or request.args.get("type") is not None:
        try:
            # sadece sayfa paramı yoksa sıfırla; varsa kullanıcı bilinçli geziniyordur
            if "page" not in request.args:
                page = 1
        except Exception:
            page = 1

    # Sayfalama
    pagination = base.order_by(Device.created_at.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    items = pagination.items  # geriye dönük: eskisi gibi 'items' değişkeni kalsın

    # Her cihaz için açık fault başlığını getir (liste görünümünde kullanmak istersen)
    open_faults = Fault.query.filter_by(is_open=True).all()
    open_fault_title = {f.device_id: f.title for f in open_faults}
    open_fault_id = {f.device_id: f.id for f in open_faults}

    # modalda kullanmak için listeler (Parametreler sayfasından)
    brands = get_param_list("brands")
    types  = get_param_list("types")
    oslist = get_param_list("os_list")

    # mevcut cihaz select’i için
    devices = Device.query.order_by(Device.brand.asc(), Device.serial.asc()).all()

    return render_template(
        "faults.html",
        items=items,               # device listesi
        pagination=pagination,     # sayfalama
        devices=devices,
        brands=brands,
        types=types,
        oslist=oslist,
        q=q,
        per_page=per_page,
        open_fault_title=open_fault_title,
        open_fault_id=open_fault_id,
        selected_brand=brand_sel,
        selected_type=type_sel,
    )

@app.route("/faults/add", methods=["POST"])
def faults_add():
    """Arıza ekleme: mevcut cihaz için başlıklı arıza aç veya yeni arızalı cihaz oluştur."""
    mode = (request.form.get("mode") or "existing").strip()

    if mode == "existing":
        # select name'i bazı şablonlarda "existingDevice" olabilir; ikisini de dene
        device_id = request.form.get("device_id", type=int) or request.form.get("existingDevice", type=int)
        title = (request.form.get("fault_title") or "").strip()
        if not device_id or not title:
            flash("Cihaz ve arıza başlığı zorunludur.", "error")
            return redirect(url_for("faults"))

        dev = Device.query.get_or_404(device_id)
        dev.is_faulty = True

        # Halihazırda açık arıza yoksa yeni kaydı aç
        open_fault = Fault.query.filter_by(device_id=device_id, is_open=True).first()
        if not open_fault:
            db.session.add(Fault(device_id=device_id, title=title, is_open=True))
        else:
            # Açık kayıt varsa başlığı güncellemek istersen:
            open_fault.title = title

        db.session.commit()
        flash("Arıza kaydı açıldı.", "ok")
        return redirect(url_for("faults"))

    # Yeni cihaz oluştur modu
    brand = (request.form.get("brand") or "").strip()
    dtype = (request.form.get("type") or "").strip()
    serial = (request.form.get("serial") or "").strip()
    home_code = (request.form.get("home_code") or "").strip()
    os_name = (request.form.get("os") or "").strip()
    quantity = request.form.get("quantity", type=int, default=1)
    title = (request.form.get("fault_title_new") or "").strip()

    if not (brand and dtype and serial and title):
        flash("Marka, Tip, Seri No ve Arıza başlığı zorunludur.", "error")
        return redirect(url_for("faults"))

    if Device.query.filter_by(serial=serial).first():
        flash("Bu seri numarası zaten kayıtlı.", "error")
        return redirect(url_for("faults"))

    dev = Device(
        brand=brand, type=dtype, serial=serial, home_code=home_code,
        os=os_name, quantity=quantity, is_faulty=True
    )
    db.session.add(dev)
    db.session.flush()  # dev.id almak için

    db.session.add(Fault(device_id=dev.id, title=title, is_open=True))
    db.session.commit()

    flash("Yeni arızalı cihaz depoya eklendi ve arıza kaydı açıldı.", "ok")
    return redirect(url_for("faults"))

@app.route("/faults/close/<int:id>", methods=["POST"])
def faults_close(id):
    f = Fault.query.get_or_404(id)
    f.is_open = False
    f.resolved_at = datetime.utcnow()
    # Cihazı normal yap
    if f.device:
        f.device.is_faulty = False
    db.session.commit()
    flash("Arıza kapatıldı.", "ok")
    return redirect(url_for("faults"))

# -------- Service (Servis) --------
@app.route("/service")
def service():
    q = (request.args.get("q") or "").strip()
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 10, type=int)

    base = ServiceRecord.query.join(Device, ServiceRecord.device_id == Device.id)

    if q:
        like = f"%{q}%"
        # Sadece seri değil; marka/tip/home_code/os’ta da ara
        base = base.filter(db.or_(
            Device.serial.ilike(like),
            Device.brand.ilike(like),
            Device.type.ilike(like),
            Device.home_code.ilike(like),
            Device.os.ilike(like),
        ))
        page = 1

    pagination = base.order_by(
        ServiceRecord.service_date.desc(), ServiceRecord.id.desc()
    ).paginate(page=page, per_page=per_page, error_out=False)

    devices = Device.query.order_by(Device.brand.asc()).all()
    brands = get_param_list("brands")
    types  = get_param_list("types")
    oslist = get_param_list("os_list")

    return render_template(
        "service.html",
        pagination=pagination,
        devices=devices,
        q=q,
        brands=brands,
        types=types,
        oslist=oslist,
        per_page=per_page,
    )

@app.route("/service/add", methods=["POST"])
def service_add():
    device_id = request.form.get("device_id", type=int)
    service_date = parse_date(request.form.get("service_date", ""))
    status = request.form.get("status", "Beklemede")
    description = request.form.get("description", "").strip()
    cost = request.form.get("cost", type=float, default=0.0)
    return_date = parse_date_optional(request.form.get("return_date"))

    if not device_id:
        flash("Cihaz seçilmelidir.", "error")
        return redirect(url_for("service"))

    # Aynı cihaz için açık (çözülmemiş) servis kaydı varsa EKLEME!
    open_exists = (
        ServiceRecord.query
        .filter(
            ServiceRecord.device_id == device_id,
            ServiceRecord.status.in_(list(OPEN_STATUSES))
        )
        .count() > 0
    )
    if open_exists and not is_resolved_status(status):
        flash("Bu cihaz için zaten açık bir servis kaydı var. Tekrar gönderilemez.", "error")
        return redirect(request.referrer or url_for("faults"))

    rec = ServiceRecord(
        device_id=device_id,
        service_date=service_date,
        status=status,
        description=description,
        cost=cost
    )

    # Çözülmüş durumsa dönüş tarihi ata, cihazı arızadan çıkar ve açık arızayı kapat
    if is_resolved_status(status):
        rec.return_date = return_date or service_date or datetime.utcnow().date()
        dev = Device.query.get(device_id)
        if dev:
            dev.is_faulty = False
        for f in Fault.query.filter_by(device_id=device_id, is_open=True).all():
            f.is_open = False
            f.resolved_at = datetime.utcnow()

    db.session.add(rec)
    db.session.commit()
    flash("Servis kaydı eklendi.", "ok")
    return redirect(url_for("service"))

@app.route("/service/update/<int:id>", methods=["POST"])
def service_update(id):
    rec = ServiceRecord.query.get_or_404(id)
    rec.device_id = request.form.get("device_id", type=int, default=rec.device_id)
    rec.service_date = parse_date(request.form.get("service_date", str(rec.service_date)))
    new_status = request.form.get("status", rec.status)
    rec.description = request.form.get("description", rec.description).strip()
    rec.cost = request.form.get("cost", type=float, default=rec.cost)

    # Manuel dönüş tarihi gönderildiyse al
    posted_return = parse_date_optional(request.form.get("return_date"))
    if posted_return:
        rec.return_date = posted_return

    rec.status = new_status

    # Güncellemede "Tamir/Teslim" olursa dönüş tarihi yoksa bugünü yaz, cihazı arızadan çıkar, açık arızayı kapat
    dev = Device.query.get(rec.device_id)
    if is_resolved_status(new_status):
        if not rec.return_date:
            rec.return_date = datetime.utcnow().date()
        if dev:
            dev.is_faulty = False
        for f in Fault.query.filter_by(device_id=rec.device_id, is_open=True).all():
            f.is_open = False
            f.resolved_at = datetime.utcnow()
    # çözülmemişe dönülürse dönüş tarihini koruyoruz (istersen temizleyebilirsin)

    db.session.commit()
    flash("Servis kaydı güncellendi.", "ok")
    return redirect(url_for("service"))

@app.route("/service/delete/<int:id>", methods=["POST"])
def service_delete(id):
    rec = ServiceRecord.query.get_or_404(id)
    db.session.delete(rec)
    db.session.commit()
    flash("Servis kaydı silindi.", "ok")
    return redirect(url_for("service"))

# -------- Params (Parametreler) --------
@app.route("/params")
def params():
    # Üç listeyi oku
    brands = get_param_list("brands")
    types  = get_param_list("types")
    oslist = get_param_list("os_list")

    # Kullanım sayıları (silme kontrolü için)
    brand_counts = {v: Device.query.filter_by(brand=v).count() for v in brands}
    type_counts  = {v: Device.query.filter_by(type=v).count()  for v in types}
    os_counts    = {v: Device.query.filter_by(os=v).count()    for v in oslist}

    return render_template(
        "params.html",
        brands=brands, types=types, oslist=oslist,
        brand_counts=brand_counts, type_counts=type_counts, os_counts=os_counts
    )

# Item ekle (brands/types/os_list)
@app.route("/params/items/<key>/add", methods=["POST"])
def params_item_add(key):
    if key not in ("brands", "types", "os_list"):
        flash("Geçersiz anahtar.", "error")
        return redirect(url_for("params"))

    val = (request.form.get("value") or "").strip()
    if not val:
        flash("Değer boş olamaz.", "error")
        return redirect(url_for("params"))

    current = get_param_list(key)
    if val in current:
        flash("Bu değer zaten var.", "error")
        return redirect(url_for("params"))

    current.append(val)
    set_param_list(key, current)
    flash("Eklendi.", "ok")
    return redirect(url_for("params"))

# Item sil (kullanılıyorsa engelle)
@app.route("/params/items/<key>/delete", methods=["POST"])
def params_item_delete(key):
    if key not in ("brands", "types", "os_list"):
        flash("Geçersiz anahtar.", "error")
        return redirect(url_for("params"))

    val = (request.form.get("value") or "").strip()
    if not val:
        flash("Değer bulunamadı.", "error")
        return redirect(url_for("params"))

    in_use = 0
    if key == "brands":
        in_use = Device.query.filter_by(brand=val).count()
    elif key == "types":
        in_use = Device.query.filter_by(type=val).count()
    elif key == "os_list":
        in_use = Device.query.filter_by(os=val).count()

    if in_use > 0:
        flash(f"Silinemedi: Depoda '{val}' değerini kullanan {in_use} cihaz var.", "error")
        return redirect(url_for("params"))

    current = [x for x in get_param_list(key) if x != val]
    set_param_list(key, current)
    flash("Silindi.", "ok")
    return redirect(url_for("params"))


# --- kullanımda liste ---
@app.route("/inuse")
def inuse():
    q = (request.args.get("q") or "").strip()
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 10, type=int)

    base = Usage.query.join(Device, Usage.device_id == Device.id).filter(Usage.is_active.is_(True))
    if q:
        like = f"%{q}%"
        base = base.filter(db.or_(
            Usage.user_name.ilike(like),
            Device.brand.ilike(like),
            Device.type.ilike(like),
            Device.serial.ilike(like),
            Device.home_code.ilike(like),
            Device.os.ilike(like),
        ))
        page = 1

    pagination = base.order_by(Usage.assigned_date.desc(), Usage.id.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    open_service_ids = {
    row[0]
    for row in (
        db.session.query(ServiceRecord.device_id)
        .filter(ServiceRecord.status.in_(list(OPEN_STATUSES)))
        .distinct()
        .all()
    )
    }
    active_usage_ids = {
        row[0]
        for row in (
            db.session.query(Usage.device_id)
            .filter(Usage.is_active.is_(True))
            .distinct()
            .all()
        )
    }

    devices = (
        Device.query
        .filter(Device.is_faulty.is_(False))
        .filter(~Device.id.in_(open_service_ids))
        .filter(~Device.id.in_(active_usage_ids))
        .order_by(Device.brand.asc(), Device.serial.asc())
        .all()
    )
    return render_template("inuse.html", pagination=pagination, devices=devices, q=q, per_page=per_page)

# --- kullanıma ver ---
@app.route("/usage/add", methods=["POST"])
def usage_add():
    device_id = request.form.get("device_id", type=int)
    user_name = (request.form.get("user_name") or "").strip()
    assigned_date = parse_date_optional(request.form.get("assigned_date")) or datetime.utcnow().date()

    if not device_id or not user_name:
        flash("Cihaz ve kullanıcı adı zorunludur.", "error")
        return redirect(url_for("inuse"))

    # hâlihazırda aktif kullanım varsa engelle
    if Usage.query.filter_by(device_id=device_id, is_active=True).count() > 0:
        flash("Bu cihaz zaten kullanımda.", "error")
        return redirect(url_for("inuse"))

    usage = Usage(device_id=device_id, user_name=user_name, assigned_date=assigned_date, is_active=True)
    db.session.add(usage)
    db.session.commit()
    flash("Cihaz kullanıma verildi.", "ok")
    return redirect(url_for("inuse"))

# --- İADE: sağlam veya arızalı ---
@app.route("/usage/return/<int:id>", methods=["POST"])
def usage_return(id):
    u = Usage.query.get_or_404(id)
    if not u.is_active:
        flash("Kayıt zaten kapatılmış.", "info")
        return redirect(url_for("inuse"))

    condition = (request.form.get("return_condition") or "ok").lower()  # "ok" | "faulty"
    u.is_active = False
    u.return_date = datetime.utcnow().date()

    dev = u.device
    if condition == "faulty":
        # Arızalı iade: cihazı arızalı yap + açık fault aç
        dev.is_faulty = True
        _open_fault_if_missing(dev.id, f"Kullanıcıdan arızalı iade ({dev.serial})")
        flash("Cihaz arızalı olarak depoya iade alındı.", "ok")
    else:
        # Sağlam iade: cihazı normal yap
        dev.is_faulty = False
        flash("Cihaz sağlam olarak depoya iade alındı.", "ok")

    db.session.commit()
    return redirect(url_for("inuse"))

# --- (opsiyonel) kullanım kaydını sil ---
@app.route("/usage/delete/<int:id>", methods=["POST"])
def usage_delete(id):
    u = Usage.query.get_or_404(id)
    db.session.delete(u)
    db.session.commit()
    flash("Kullanım kaydı silindi.", "ok")
    return redirect(url_for("inuse"))

# -------- REPORTS --------
@app.route("/reports")
def reports():
    # Açık servis kaydı olan cihazlar (cihaz bazında tekil)
    open_q = (
        db.session.query(ServiceRecord.device_id)
        .filter(ServiceRecord.status.in_(list(OPEN_STATUSES)))
        .distinct()
    )
    open_device_ids = {row[0] for row in open_q.all()}

    # Aktif kullanımda olan cihazlar
    in_use_device_ids = {
        row[0]
        for row in (
            db.session.query(Usage.device_id)
            .filter(Usage.return_date.is_(None))
            .distinct()
            .all()
        )
    }

    all_devices = Device.query.all()

    # ---- sayımlar ----
    faulty_count  = sum(1 for d in all_devices if d.is_faulty)
    service_count = sum(1 for d in all_devices if d.id in open_device_ids)
    inuse_count   = sum(1 for d in all_devices if d.id in in_use_device_ids)
    healthy_count = sum(
        1 for d in all_devices
        if (not d.is_faulty) and (d.id not in open_device_ids) and (d.id not in in_use_device_ids)
    )

    # ---- gruplamalar yardımcı ----
    def count_by(attr, iterable):
        from collections import defaultdict
        c = defaultdict(int)
        for x in iterable:
            key = getattr(x, attr) or "—"
            c[key] += 1
        labels = list(c.keys())
        values = [c[k] for k in labels]
        return labels, values

    # listeler
    healthy_devices = [d for d in all_devices
                       if (not d.is_faulty) and (d.id not in open_device_ids) and (d.id not in in_use_device_ids)]
    faulty_devices  = [d for d in all_devices if d.is_faulty]
    service_devices = Device.query.filter(Device.id.in_(open_device_ids)).all()
    inuse_devices   = Device.query.filter(Device.id.in_(in_use_device_ids)).all()

    # marka/tipe göre
    healthy_brand_labels, healthy_brand_values = count_by("brand", healthy_devices)
    healthy_type_labels,  healthy_type_values  = count_by("type",  healthy_devices)

    faulty_brand_labels,  faulty_brand_values  = count_by("brand", faulty_devices)
    faulty_type_labels,   faulty_type_values   = count_by("type",  faulty_devices)

    service_brand_labels, service_brand_values = count_by("brand", service_devices)
    service_type_labels,  service_type_values  = count_by("type",  service_devices)

    inuse_brand_labels,   inuse_brand_values   = count_by("brand", inuse_devices)
    inuse_type_labels,    inuse_type_values    = count_by("type",  inuse_devices)

    # Kullanımdaki kayıtlar (kullanıcı dağılımı)
    active_usages = (
        Usage.query.filter(Usage.return_date.is_(None))
        .join(Device, Usage.device_id == Device.id)
        .all()
    )
    from collections import defaultdict
    user_count = defaultdict(int)
    for u in active_usages:
        user_count[(u.user_name or "—")] += 1
    inuse_user_labels  = list(user_count.keys())
    inuse_user_values  = [user_count[k] for k in inuse_user_labels]

    return render_template(
        "reports.html",
        # genel durum + yeni "inuse_count"
        healthy_count=healthy_count,
        faulty_count=faulty_count,
        service_count=service_count,
        inuse_count=inuse_count,

        # marka
        healthy_brand_labels=healthy_brand_labels, healthy_brand_values=healthy_brand_values,
        faulty_brand_labels=faulty_brand_labels,   faulty_brand_values=faulty_brand_values,
        service_brand_labels=service_brand_labels, service_brand_values=service_brand_values,
        inuse_brand_labels=inuse_brand_labels,     inuse_brand_values=inuse_brand_values,

        # tip
        healthy_type_labels=healthy_type_labels, healthy_type_values=healthy_type_values,
        faulty_type_labels=faulty_type_labels,   faulty_type_values=faulty_type_values,
        service_type_labels=service_type_labels, service_type_values=service_type_values,
        inuse_type_labels=inuse_type_labels,     inuse_type_values=inuse_type_values,

        # kullanıcı
        inuse_user_labels=inuse_user_labels,
        inuse_user_values=inuse_user_values,
    )

# ---------------------------
# Yardımcı: SQLAlchemy objelerini DataFrame'e dönüştür
# ---------------------------
def df_from_query(query, columns=None, mapper=None):
    """
    query: SQLAlchemy query (Model.query.all() gibi)
    columns: İstenirse belirli kolon isimleri listesi
    mapper: Her kayıt için dict döndüren fonksiyon (özelleştirme)
    """
    rows = query.all() if hasattr(query, "all") else list(query)
    if mapper:
        data = [mapper(r) for r in rows]
        return pd.DataFrame(data)
    else:
        # Basit yolla __dict__ (SQLAlchemy ek alanlarını filtreleriz)
        def row_to_dict(obj):
            d = {}
            for k in obj.__dict__.keys():
                if not k.startswith("_sa_"):
                    v = getattr(obj, k)
                    # datetime biçimleme
                    if isinstance(v, datetime):
                        v = v.strftime("%Y-%m-%d %H:%M:%S")
                    d[k] = v
            return d
        data = [row_to_dict(r) for r in rows]
        if columns:
            # Kolon sırasını sabitle
            return pd.DataFrame(data)[columns]
        return pd.DataFrame(data)

# ---------------------------
# Excel çıktısı
# ---------------------------
@app.route("/reports/download/excel")
def reports_download_excel():
    import io
    from openpyxl.utils import get_column_letter
    import pandas as pd
    from collections import defaultdict

    # ---------- Sorgular ----------
    devices_q  = Device.query.order_by(Device.created_at.desc())
    services_q = ServiceRecord.query.order_by(ServiceRecord.service_date.desc(), ServiceRecord.id.desc())
    faults_q   = Fault.query.order_by(Fault.created_at.desc())

    # Kullanımda (iade edilmemiş) cihazlar
    inuse_q = (
        Usage.query
        .filter(Usage.return_date.is_(None))
        .order_by(Usage.assigned_date.desc(), Usage.id.desc())
    )

    # ---------- İlişkisiz güvenli erişim için cihaz cache'i ----------
    fault_rows   = faults_q.all()
    service_rows = services_q.all()
    inuse_rows   = inuse_q.all()

    dev_ids = set()
    dev_ids.update([getattr(f, "device_id", None) for f in fault_rows   if getattr(f, "device_id", None)])
    dev_ids.update([getattr(s, "device_id", None) for s in service_rows if getattr(s, "device_id", None)])
    dev_ids.update([getattr(u, "device_id", None) for u in inuse_rows   if getattr(u, "device_id", None)])

    device_cache = {d.id: d for d in Device.query.filter(Device.id.in_(dev_ids)).all()}
    def get_dev(did):
        return device_cache.get(did)

    # ---------- DataFrame yardımcıları ----------
    def device_map(d: Device):
        return {
            "ID": d.id,
            "Marka": d.brand,
            "Tip": d.type,
            "Seri No": d.serial,
            "Home Code": d.home_code,
            "OS": d.os,
            "Adet": d.quantity,
            "Arızalı mı?": "Evet" if d.is_faulty else "Hayır",
            "Depo Tarihi": d.created_at.strftime("%Y-%m-%d") if d.created_at else None,
        }

    def service_map(s: ServiceRecord):
        d = get_dev(getattr(s, "device_id", None))
        return {
            "Kayıt ID": s.id,
            "Cihaz ID": getattr(s, "device_id", None),
            "Marka": getattr(d, "brand", None),
            "Tip": getattr(d, "type", None),
            "Seri No": getattr(d, "serial", None),
            "Gönderim Tarihi": s.service_date.strftime("%Y-%m-%d") if getattr(s, "service_date", None) else None,
            "Durum": getattr(s, "status", None),
            "Açıklama": getattr(s, "description", None),
            "Maliyet": getattr(s, "cost", None),
            "Dönüş Tarihi": s.return_date.strftime("%Y-%m-%d") if getattr(s, "return_date", None) else None,
            "Oluşturma": s.created_at.strftime("%Y-%m-%d") if getattr(s, "created_at", None) else None,
        }

    def fault_map(f: Fault):
        d = get_dev(getattr(f, "device_id", None))
        return {
            "Kayıt ID": f.id,
            "Cihaz ID": getattr(f, "device_id", None),
            "Marka": getattr(d, "brand", None),
            "Tip": getattr(d, "type", None),
            "Seri No": getattr(d, "serial", None),
            "Başlık": getattr(f, "title", None),
            "Açık mı?": "Evet" if getattr(f, "is_open", False) else "Hayır",
            "Oluşturma": f.created_at.strftime("%Y-%m-%d") if getattr(f, "created_at", None) else None,
            "Çözüm Tarihi": f.resolved_at.strftime("%Y-%m-%d %H:%M:%S") if getattr(f, "resolved_at", None) else None,
        }

    def usage_map(u: Usage):
        d = get_dev(getattr(u, "device_id", None))
        return {
            "Kullanım ID": u.id,
            "Kullanıcı": getattr(u, "user_name", None),
            "Veriliş Tarihi": u.assigned_date.strftime("%Y-%m-%d") if getattr(u, "assigned_date", None) else None,
            "Cihaz ID": getattr(u, "device_id", None),
            "Marka": getattr(d, "brand", None),
            "Tip": getattr(d, "type", None),
            "Seri No": getattr(d, "serial", None),
            "Home Code": getattr(d, "home_code", None),
            "OS": getattr(d, "os", None),
        }

    def to_df(rows, mapper):
        rows = rows if isinstance(rows, list) else rows.all()
        return pd.DataFrame([mapper(r) for r in rows])

    # ---------- DataFrame'ler ----------
    df_devices  = to_df(devices_q, device_map)
    df_services = to_df(service_rows, service_map)
    df_faults   = to_df(fault_rows,   fault_map)
    df_inuse    = to_df(inuse_rows,   usage_map)

    # Sağlam / Servis / Arızalı kümeleri
    open_service_device_ids = {
        row[0]
        for row in (
            db.session.query(ServiceRecord.device_id)
            .filter(ServiceRecord.status.in_(list(OPEN_STATUSES)))
            .distinct()
            .all()
        )
    }
    all_devices = devices_q.all()
    healthy = [d for d in all_devices if (not d.is_faulty) and (d.id not in open_service_device_ids)]
    faulty  = [d for d in all_devices if d.is_faulty]
    service = [d for d in all_devices if d.id in open_service_device_ids]
    inuse_devices = [get_dev(u.device_id) for u in inuse_rows if get_dev(u.device_id)]

    # Özet
    df_summary = pd.DataFrame(
        [
            {"Kategori": "Kullanımda",       "Adet": len(inuse_devices)},
            {"Kategori": "Serviste",         "Adet": len(service)},
            {"Kategori": "Arızalı Depoda",   "Adet": len(faulty)},
            {"Kategori": "Sağlam Depoda",    "Adet": len(healthy)},
            {"Kategori": "Toplam Cihaz",     "Adet": len(all_devices)},
        ]
    )

    # Dağılım yardımcıları
    def count_df(items, attr, colname):
        c = defaultdict(int)
        for x in items:
            c[getattr(x, attr) or "—"] += 1
        return pd.DataFrame({colname: list(c.keys()), "Adet": list(c.values())})

    # Markaya göre
    df_brand_healthy = count_df(healthy, "brand", "Marka")
    df_brand_faulty  = count_df(faulty,  "brand", "Marka")
    df_brand_service = count_df(service, "brand", "Marka")
    df_brand_inuse   = count_df(inuse_devices, "brand", "Marka")

    # Tip'e göre
    df_type_healthy  = count_df(healthy, "type", "Tip")
    df_type_faulty   = count_df(faulty,  "type", "Tip")
    df_type_service  = count_df(service, "type", "Tip")
    df_type_inuse    = count_df(inuse_devices, "type", "Tip")

    # ---------- Excel yaz ----------
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Ana sayfalar
        df_devices.to_excel(writer,  index=False, sheet_name="Cihazlar")
        df_services.to_excel(writer, index=False, sheet_name="Servis Kayıtları")
        df_faults.to_excel(writer,   index=False, sheet_name="Arıza Kayıtları")
        df_inuse.to_excel(writer,    index=False, sheet_name="Kullanımda")

        # Özet
        df_summary.to_excel(writer, index=False, sheet_name="Özet")

        # Dağılım sayfaları (Marka)
        df_brand_healthy.to_excel(writer, index=False, sheet_name="Marka-Sağlam")
        df_brand_faulty.to_excel(writer,  index=False, sheet_name="Marka-Arızalı")
        df_brand_service.to_excel(writer, index=False, sheet_name="Marka-Serviste")
        df_brand_inuse.to_excel(writer,   index=False, sheet_name="Marka-Kullanımda")

        # Dağılım sayfaları (Tip)
        df_type_healthy.to_excel(writer, index=False, sheet_name="Tip-Sağlam")
        df_type_faulty.to_excel(writer,  index=False, sheet_name="Tip-Arızalı")
        df_type_service.to_excel(writer, index=False, sheet_name="Tip-Serviste")
        df_type_inuse.to_excel(writer,   index=False, sheet_name="Tip-Kullanımda")

        # Auto-fit sütun genişlikleri
        for sheet_name in writer.book.sheetnames:
            ws = writer.book[sheet_name]
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        max_len = max(max_len, len(str(cell.value)) if cell.value is not None else 0)
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    output.seek(0)
    fname = f"Rapor_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(output, download_name=fname, as_attachment=True)


def _register_fonts():
    """Montserrat varsa kaydeder; hata olursa sessizce Helvetica kullanılır."""
    try:
        if os.path.isfile(MONTS_REG):
            pdfmetrics.registerFont(TTFont("Montserrat", MONTS_REG))
        if os.path.isfile(MONTS_BOLD):
            pdfmetrics.registerFont(TTFont("Montserrat-Bold", MONTS_BOLD))
    except Exception as e:
        print(f"[PDF] Font register warning: {e}")

def _font_available(name: str) -> bool:
    try:
        pdfmetrics.getFont(name)
        return True
    except Exception:
        return False

def _pick_font(bold: bool = False) -> str:
    if bold and _font_available("Montserrat-Bold"):
        return "Montserrat-Bold"
    if (not bold) and _font_available("Montserrat"):
        return "Montserrat"
    return "Helvetica-Bold" if bold else "Helvetica"

def _make_styles():
    """Kurumsal görünümlü paragraflar."""
    base = getSampleStyleSheet()
    body_font = _pick_font(False)
    bold_font = _pick_font(True)
    styles = {
        "Title": ParagraphStyle(
            "Title",
            parent=base["Title"],
            fontName=bold_font, fontSize=18, leading=22, spaceAfter=6, textColor=colors.HexColor("#111111")
        ),
        "Subtle": ParagraphStyle(
            "Subtle",
            parent=base["BodyText"],
            fontName=body_font, fontSize=9, leading=12, textColor=colors.HexColor("#475467")
        ),
        "H2": ParagraphStyle(
            "H2",
            parent=base["Heading2"],
            fontName=bold_font, fontSize=13, leading=16, spaceBefore=10, spaceAfter=6
        ),
        "H3": ParagraphStyle(
            "H3",
            parent=base["Heading3"],
            fontName=bold_font, fontSize=11.5, leading=14, spaceBefore=8, spaceAfter=4
        ),
        "Body": ParagraphStyle(
            "Body",
            parent=base["BodyText"],
            fontName=body_font, fontSize=10, leading=13
        ),
    }
    return styles

def _on_page(canvas, doc):
    """Her sayfada başlık/altbilgi çizimi."""
    W, H = landscape(A4)
    canvas.saveState()
    # Üst şerit
    canvas.setStrokeColorRGB(0, 0, 0)
    canvas.setLineWidth(0.6)
    canvas.line(24, H - 36, W - 24, H - 36)

    # Sol üstte ürün adı
    canvas.setFont(_pick_font(True), 11)
    canvas.drawString(24, H - 28, "METTA IT — Rapor")

    # Sağ üstte tarih
    canvas.setFont(_pick_font(False), 9)
    canvas.drawRightString(W - 24, H - 28, datetime.now().strftime("%d.%m.%Y %H:%M"))

    # Alt şerit + sayfa no
    canvas.setLineWidth(0.4)
    canvas.line(24, 28, W - 24, 28)
    canvas.setFont(_pick_font(False), 9)
    canvas.drawRightString(W - 24, 16, f"Sayfa {doc.page}")
    canvas.restoreState()

def _table(style_header_bg="#111111", zebra=False):
    """Tablo style helper."""
    base = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(style_header_bg)),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.whitesmoke),
        ("FONTNAME",   (0, 0), (-1, 0), _pick_font(True)),
        ("FONTSIZE",   (0, 0), (-1, 0), 10),
        ("ALIGN",      (0, 0), (-1, 0), "CENTER"),
        ("GRID",       (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ("FONTSIZE",   (0, 1), (-1, -1), 9),
        ("FONTNAME",   (0, 1), (-1, -1), _pick_font(False)),
        ("LEFTPADDING",(0, 0), (-1, -1), 6),
        ("RIGHTPADDING",(0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 6),
    ]
    if zebra:
        base.append(("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#ffffff"), colors.HexColor("#F7F7F7")]))
    return TableStyle(base)

def _count_by_attr(devices, attr):
    from collections import defaultdict
    c = defaultdict(int)
    for d in devices:
        key = getattr(d, attr) or "—"
        c[key] += 1
    # çoktan aza, sonra alfabetik
    return sorted(c.items(), key=lambda kv: (-kv[1], kv[0]))

# ---------------------------
# PDF çıktısı (özet + tablo)
# ---------------------------
@app.route("/reports/download/pdf")
def reports_download_pdf():
    # 1) Fontları kaydet + stiller
    _register_fonts()
    styles = _make_styles()

    # 2) Verileri hesapla (rapor sayfasıyla tutarlı)
    open_service_device_ids = {
        row[0]
        for row in (
            db.session.query(ServiceRecord.device_id)
            .filter(ServiceRecord.status.in_(list(OPEN_STATUSES)))
            .distinct()
            .all()
        )
    }
    all_devices = Device.query.all()
    healthy = [d for d in all_devices if (not d.is_faulty) and (d.id not in open_service_device_ids)]
    faulty  = [d for d in all_devices if d.is_faulty]
    service = [d for d in all_devices if d.id in open_service_device_ids]

    inuse_devices = (
        db.session.query(Device)
        .join(Usage, Usage.device_id == Device.id)
        .filter(Usage.return_date.is_(None))
        .all()
    )

    # 3) PDF belge
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=24, rightMargin=24, topMargin=48, bottomMargin=36
    )

    elems = []

    # Başlık
    title = f"Genel Rapor — {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    elems.append(Paragraph("METTA IT", styles["Title"]))
    elems.append(Paragraph(title, styles["Subtle"]))
    elems.append(Spacer(1, 10))

    # Özet
    elems.append(Paragraph("Özet", styles["H2"]))
    summary_data = [
        ["Kategori",          "Adet"],
        ["Sağlam (Depoda)",   len(healthy)],
        ["Arızalı (Depoda)",  len(faulty)],
        ["Serviste",          len(service)],
        ["Toplam",            len(all_devices)],
    ]
    t = Table(summary_data, colWidths=[240, 100])
    t.setStyle(_table(style_header_bg="#0F172A", zebra=True))
    elems.append(t)
    elems.append(Spacer(1, 14))

    # Son 25 cihaz tablosu
    elems.append(Paragraph("Son 25 Cihaz (Marka / Tip / Seri / Adet / Durum / Depo Tarihi)", styles["H3"]))
    last_devices = Device.query.order_by(Device.created_at.desc()).limit(25).all()
    ld_rows = [["Marka", "Tip", "Seri No", "Adet", "Durum", "Depo Tarihi"]]
    for d in last_devices:
        ld_rows.append([
            d.brand or "—",
            d.type or "—",
            d.serial or "—",
            d.quantity or 1,
            "Arızalı" if d.is_faulty else "Depoda",
            d.created_at.strftime("%Y-%m-%d")
        ])
    t2 = Table(ld_rows, colWidths=[150, 150, 180, 60, 80, 110])
    t2.setStyle(_table(style_header_bg="#111111", zebra=True))
    elems.append(t2)
    elems.append(Spacer(1, 12))

    # Marka kırılımları
    elems.append(Paragraph("Markaya Göre Dağılımlar", styles["H2"]))

    def add_dist_section(header, devices_list):
        elems.append(Paragraph(header, styles["H3"]))
        data = [["Marka", "Adet"]] + [[k, v] for k, v in _count_by_attr(devices_list, "brand")]
        tbl = Table(data, colWidths=[300, 80])
        tbl.setStyle(_table(style_header_bg="#1F2937", zebra=True))
        elems.append(tbl)
        elems.append(Spacer(1, 10))

    add_dist_section("Sağlam (Depoda)", healthy)
    add_dist_section("Arızalı (Depoda)", faulty)
    add_dist_section("Serviste", service)

    # Tip kırılımları
    elems.append(Paragraph("Tip Bazlı Dağılımlar", styles["H2"]))

    def add_type_section(header, devices_list):
        elems.append(Paragraph(header, styles["H3"]))
        data = [["Tip", "Adet"]] + [[k, v] for k, v in _count_by_attr(devices_list, "type")]
        tbl = Table(data, colWidths=[300, 80])
        tbl.setStyle(_table(style_header_bg="#334155", zebra=True))
        elems.append(tbl)
        elems.append(Spacer(1, 10))

    add_type_section("Sağlam (Depoda)", healthy)
    add_type_section("Arızalı (Depoda)", faulty)
    add_type_section("Serviste", service)

    # --- Kullanım bölümü (PDF) ---
    elems.append(PageBreak())
    elems.append(Paragraph("Kullanımda Olan Cihazlar", styles["H2"]))

    # Özet kutusu
    usage_summary = [
        ["Kategori", "Adet"],
        ["Kullanımda", len(inuse_devices)],
    ]
    tbl_u0 = Table(usage_summary, colWidths=[220, 80])
    tbl_u0.setStyle(_table(style_header_bg="#0F172A", zebra=True))
    elems.append(tbl_u0)
    elems.append(Spacer(1, 10))

    # Marka dağılımı
    elems.append(Paragraph("Markaya Göre Kullanım", styles["H3"]))
    u_brand = [["Marka", "Adet"]] + [[k, v] for k, v in _count_by_attr(inuse_devices, "brand")]
    tbl_u1 = Table(u_brand, colWidths=[300, 80]); tbl_u1.setStyle(_table(style_header_bg="#1F2937", zebra=True))
    elems.append(tbl_u1); elems.append(Spacer(1, 8))

    # Tip dağılımı
    elems.append(Paragraph("Tip Bazlı Kullanım", styles["H3"]))
    u_type = [["Tip", "Adet"]] + [[k, v] for k, v in _count_by_attr(inuse_devices, "type")]
    tbl_u2 = Table(u_type, colWidths=[300, 80]); tbl_u2.setStyle(_table(style_header_bg="#334155", zebra=True))
    elems.append(tbl_u2); elems.append(Spacer(1, 8))

    # Son 20 kullanım (kullanıcı bazlı)
    elems.append(Paragraph("Son 20 Kullanım Kaydı", styles["H3"]))
    last_usages = (
        Usage.query.filter(Usage.return_date.is_(None))
        .order_by(Usage.assigned_date.desc(), Usage.id.desc())
        .limit(20).all()
    )
    rows = [["Kullanıcı","Marka","Tip","Seri No","Veriliş"]]
    for u in last_usages:
        d = u.device
        rows.append([u.user_name or "—", d.brand or "—", d.type or "—", d.serial or "—",
                    u.assigned_date.strftime("%Y-%m-%d") if u.assigned_date else "—"])
    tbl_u3 = Table(rows, colWidths=[140,120,120,160,80]); tbl_u3.setStyle(_table(style_header_bg="#111111", zebra=True))
    elems.append(tbl_u3)

    # Dipnot
    elems.append(Spacer(1, 6))
    elems.append(Paragraph(
        "Not: Bu PDF, METTA IT için oluşturulmuştur.",
        styles["Subtle"]
    ))

    # 4) Derle — her sayfaya şablon çiz
    doc.build(elems, onFirstPage=_on_page, onLaterPages=_on_page)

    buf.seek(0)
    fname = f"Rapor_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    return send_file(buf, download_name=fname, as_attachment=True)
# ----------------------------
if __name__ == "__main__":
    app.run(debug=True,)
